var_lst=['67','87','23','55','98']
from openpyxl import load_workbook
import xlsxwriter
from xlsxwriter.utility import xl_rowcol_to_cell
import xlrd
myPath= r'C:\Users\hp world\Desktop\hello.xlsx'
wbk = xlsxwriter.Workbook(myPath)
wks = wbk.add_worksheet()
wks.write(0,0,"list")
for i in range(1,100):
    wks.write(i,0,i)

wbk.close()
workbook=load_workbook(filename=myPath)
sheet=workbook.active
for sh in xlrd.open_workbook(myPath).sheets():  
    for row in range(sh.nrows):
        for col in range(sh.ncols):
            for v in var_lst:
                if int(v) == sh.cell(row, col).value:
                    sheet[xl_rowcol_to_cell(row,col)]=int(v)*int(v)                 
workbook.save(filename=r"C:\Users\hp world\Desktop/hello.xlsx")               
